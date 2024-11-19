import pandas as pd
import os
from datetime import datetime, timedelta
import openpyxl
from openpyxl.styles import PatternFill

# Dates information
classes_taken_dates = ["06/08/2024", "13/08/2024", "20/08/2024", "27/08/2024",
                       "03/09/2024", "17/09/2024", "01/10/2024"]
classes_missed_dates = ["10/09/2024"]
exams_dates = ["24/09/2024"]
all_dates = sorted(classes_taken_dates + classes_missed_dates + exams_dates, key=lambda x: datetime.strptime(x, '%d/%m/%Y'))  # Sort all dates in chronological order
class_timing = "18:00 - 20:00"

# Helper function to check if a timestamp falls within the lecture period (6:00 PM - 8:00 PM)
def is_present(attendance_time, lecture_start, lecture_end):
    attendance_time = datetime.strptime(attendance_time, '%d/%m/%Y %H:%M:%S')  # Adjust for seconds
    return lecture_start <= attendance_time <= lecture_end

# Determine the directory of the script
current_directory = os.path.dirname(__file__)

# Read student list (Extract Roll Number Only)
with open(os.path.join(current_directory, 'stud_list.txt'), 'r') as file:
    students = [line.split()[0] for line in file.read().splitlines()]  # Extracting only Roll Number

# Read the input CSV properly
attendance_df = pd.read_csv(os.path.join(current_directory, 'input_attendance.csv'), names=['Timestamp', 'Roll'], header=0)

# Check for NaN values and handle them
attendance_df['Roll'] = attendance_df['Roll'].fillna('')  # Fill NaN with empty strings

# Extract the roll number and name only if Roll is not empty
attendance_df['Roll'] = attendance_df['Roll'].apply(lambda x: x.split()[0] if x else '')  # Extract roll number

# Generate the full path for the output Excel file
output_excel = os.path.join(current_directory, 'output_excel.xlsx')

# Initialize the Excel writer
writer = pd.ExcelWriter(output_excel, engine='openpyxl')

# Initialize attendance summary and proxy summary
attendance_summary = []
proxy_summary = []

# Process attendance for each date (classes, missed classes, and exams)
for date in all_dates:
    lecture_start = datetime.strptime(f'{date} 18:00', '%d/%m/%Y %H:%M')
    lecture_end = lecture_start + timedelta(hours=2)  # From 6:00 PM to 8:00 PM

    # Track attendance and proxy for each student
    daily_attendance = {}
    daily_proxy = {}

    for student in students:
        # Filter student attendance records for the given date
        student_records = attendance_df[(attendance_df['Roll'] == student) &
                                        (attendance_df['Timestamp'].str.startswith(date))]

        # Count the number of valid attendance records within the lecture period
        present_count = student_records['Timestamp'].apply(lambda x: is_present(x, lecture_start, lecture_end)).sum()

        # Set the attendance count based on actual attendance records
        attendance_status = present_count  # Keep the exact number of times attendance was marked

        # If the student marked attendance more than 2 times, increment the proxy count
        proxy_count = present_count - 2 if present_count > 2 else 0

        daily_attendance[student] = attendance_status
        daily_proxy[student] = proxy_count

    attendance_summary.append(daily_attendance)
    proxy_summary.append(daily_proxy)

# Convert the summaries to DataFrames
attendance_df = pd.DataFrame(attendance_summary, index=all_dates)
proxy_df = pd.DataFrame(proxy_summary, index=all_dates)

# Transpose DataFrames so students are in rows and dates are in columns
attendance_df = attendance_df.T
proxy_df = proxy_df.T

# Add summary columns for total attendance counts, total attendance marked, total allowed, and proxy counts
attendance_df['Total Count'] = attendance_df.count(axis=1)  # Count of dates with attendance marks
attendance_df['Total Attendance Marked'] = attendance_df.iloc[:, :-1].sum(axis=1)  # Sum attendance marks
attendance_df['Total Attendance Allowed'] = len(classes_taken_dates) * 2  # Total allowed attendance
attendance_df['Total Proxy'] = proxy_df.sum(axis=1)  # Total proxy count

# Write to Excel
attendance_df.to_excel(writer, sheet_name='Attendance')

# Apply conditional formatting
workbook = writer.book
sheet = workbook['Attendance']

# Define colors
red_fill = PatternFill(start_color='FF9999', end_color='FF9999', fill_type='solid')
yellow_fill = PatternFill(start_color='FFFF99', end_color='FFFF99', fill_type='solid')
green_fill = PatternFill(start_color='99FF99', end_color='99FF99', fill_type='solid')

# Apply formatting based on attendance status for classes, missed classes, and exams
for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=2, max_col=sheet.max_column - 4):  # Exclude summary columns
    for cell in row:
        if cell.value == 0:
            cell.fill = red_fill
        elif cell.value == 1:
            cell.fill = yellow_fill
        elif cell.value == 2:
            cell.fill = green_fill

# Save the file
writer.close()

print(f"Output Excel file saved at: {output_excel}")
